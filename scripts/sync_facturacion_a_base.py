"""
Sincroniza 'Facturacion clientes JVA.xlsx' (fuente de verdad, mantenida a
mano) hacia 'BASE_DE_DATOS_JVA.xlsx' (la que lee Power BI, hoja
Fact_Facturas). Reemplaza a sincronizar_facturas_nuevas.ps1 (dejó de
correr el 16-ago-2026 por un error de sintaxis y nunca hacía la parte 1).

Tres trabajos, en este orden:

1) ACTUALIZAR estado de pago de facturas ya existentes en la Base:
   corrige el flag Pagada cuando Facturacion dice "Si" y la Base dice
   "No", y completa Fecha_Real_Pago si la Base la tiene en blanco.
   Nunca pisa un estado especial ("En demanda", "Sin registrar") ni una
   Fecha_Real_Pago ya cargada en la Base.

2) INGRESAR facturas nuevas: filas de Facturacion que no existen todavía
   en la Base, con los mismos requisitos que ya aplica la consulta Power
   Query de Fact_Facturas (Fecha_Emision, Fecha_Vencimiento y Neto no
   vacíos/no cero). El Cliente_ID se resuelve por RUT contra
   Dim_Clientes:
     - RUT con un solo candidato -> se usa ese.
     - RUT con varios candidatos (casos ya documentados en
       QA_Calidad_Datos, ej. dos clientes que comparten un RUT probablemente
       mal tipeado pero son clientes reales distintos) -> solo se acepta
       si el nombre calza exacto; si no, se dejan afuera y se anotan en
       QA_Calidad_Datos para revisión manual, nunca se adivina.
     - RUT no encontrado en absoluto -> se crea un cliente nuevo en
       Dim_Clientes (Cliente_ID correlativo nuevo) con el nombre tal
       cual viene en Facturacion.

3) Antes de escribir cualquier cambio real: respaldo con timestamp de
   BASE_DE_DATOS_JVA.xlsx.

No toca el archivo de Facturacion. Pensado para correr solo (Programador
de tareas de Windows). Es seguro ejecutarlo todos los días: si no hay
cambios, no escribe ni respalda nada.
"""
import datetime
import os
import re
import shutil
import sys
from collections import defaultdict

import openpyxl

FACTURACION_PATH = r"C:\Users\ignac\Dropbox\JVA ADM\Facturas\Facturacion clientes JVA.xlsx"
BASE_PATH = r"C:\Users\ignac\Dropbox\JVA ADM\BASE_DE_DATOS_JVA.xlsx"
LOG_PATH = r"C:\Users\ignac\Dropbox\JVA ADM\Scripts\sync_facturacion_log.txt"

# Columnas 0-based en 'Facturacion clientes JVA.xlsx' (hoja FACTURAS)
F_NFAC, F_CLIENTE, F_RUT, F_FEM, F_FVENC, F_NETO, F_IVA, F_TOTAL = 2, 3, 4, 5, 6, 7, 8, 9
F_OC, F_GUIA, F_CEDIDA, F_FACTORING, F_PAGADA, F_BANCO = 10, 11, 12, 13, 14, 15

# Columnas 1-based en 'BASE_DE_DATOS_JVA.xlsx' (hoja Fact_Facturas)
B_FACTURA_ID, B_NCORR, B_NFAC, B_TIPODOC, B_CLIENTE_ID = 1, 2, 3, 4, 5
B_CLIENTE_REG, B_RUT_REG, B_FEM, B_FVENC, B_NETO = 6, 7, 8, 9, 10
B_IVA, B_TOTAL, B_OC, B_GUIA, B_CEDIDA, B_FACTORING = 11, 12, 13, 14, 15, 16
B_PAGADA, B_BANCO, B_FECHA_REAL_PAGO, B_BANCO_PAGO = 17, 18, 19, 20

# Columnas 1-based en Dim_Clientes
DC_CLIENTE_ID, DC_NOMBRE, DC_RUT, DC_N_OTS, DC_N_FACTURAS = 1, 2, 3, 4, 5


def log(msg):
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{stamp}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def norm_rut(rut):
    if not rut:
        return None
    return re.sub(r"[.\s]", "", str(rut)).upper()


# Comentarios de celda en la columna 'Pagada' del tipo:
#   "Pc: Pagado en bco chile el 21-08-2026"
#   "Pc: Pagado en scotiabank el 19-08-2026"
# Formato de fecha DD-MM-AAAA o DD-MM-AA (día-mes-año, separador '-' o '/').
_FECHA_COMENTARIO_RE = re.compile(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})")

# Ninguna factura de JVA es anterior a esto; una fecha de pago fuera de
# este rango es un error de tipeo, no un pago.
FECHA_PAGO_MIN = datetime.datetime(2015, 1, 1)
FECHA_PAGO_MARGEN_FUTURO = datetime.timedelta(days=365)


def _repara_anio(y):
    """Repara años a los que se les perdió el primer dígito al tipear.

    En la planilla hay comentarios como "Abonado en bco chile el 7-12-222":
    el año 2022 quedó como 222 porque se comió el '2' inicial. Como 222
    es >= 100, la corrección de dos dígitos no lo tocaba y datetime lo
    aceptaba como el año 222 d.C. El Dias_Atraso resultante era de
    -657.000 días, lo que reventaba cualquier promedio de DPD (el de 2022
    daba -25.849 días). 8 casos detectados el 2026-09-01: facturas
    107, 108, 109, 178, 181, 207, 208 y 263.
    """
    if y < 100:            # "26"  -> 2026
        return y + 2000
    if y < 1000:           # "222" -> 2022, "223" -> 2023
        return 2000 + (y % 100)
    return y


def fecha_pago_plausible(fecha, origen=""):
    """Descarta fechas de pago imposibles en vez de propagarlas al modelo."""
    if fecha is None:
        return None
    limite = datetime.datetime.now() + FECHA_PAGO_MARGEN_FUTURO
    if not (FECHA_PAGO_MIN <= fecha <= limite):
        log(f"QA: fecha de pago implausible descartada ({fecha:%d-%m-%Y}) {origen}")
        return None
    return fecha


def fecha_desde_comentario(texto_comentario):
    if not texto_comentario:
        return None
    m = _FECHA_COMENTARIO_RE.search(texto_comentario)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    y = _repara_anio(y)
    try:
        fecha = datetime.datetime(y, mo, d)
    except ValueError:
        return None
    return fecha_pago_plausible(fecha, f"en comentario: {texto_comentario!r}")


def estado_pago_facturacion(fr):
    """Determina si una fila de Facturacion está realmente pagada.

    El flag manual 'Pagada' (Sí/No) lo actualiza el jefe a mano y a veces
    se le olvida marcarlo aunque la factura ya esté pagada. Hay dos señales
    más confiables que ese flag, en orden de prioridad:

    1) Un COMENTARIO de celda en la columna 'Pagada' (ej. una factura con
       comentario "Pagado en banco X el DD-MM-AAAA" pero el flag todavía
       en 'No'). Es la anotación manual de quien concilia el pago, así que
       es la fuente más confiable cuando existe — 10 casos documentados en
       un barrido de la planilla completa.
    2) Una fecha real tipeada en la columna 'Banco' en vez del nombre del
       banco (una de esas mismas facturas también tenía una fecha ahí,
       distinta/menos confiable que la del comentario — por eso el
       comentario tiene prioridad si ambos están presentes).

    fr[-1] lleva el texto del comentario de la celda 'Pagada' (ver
    load_facturacion). Devuelve (pagada: bool, fecha_pago: datetime|None).
    """
    pagada_txt = (fr[F_PAGADA] or "").strip()
    fecha_comentario = fecha_desde_comentario(fr[-1])
    banco = fr[F_BANCO]
    fecha_banco = fecha_pago_plausible(
        banco if isinstance(banco, datetime.datetime) else None,
        f"en columna Banco (factura {fr[F_NFAC]})")
    fecha_pago = fecha_comentario or fecha_banco
    if fecha_pago is not None:
        return True, fecha_pago
    return pagada_txt.lower() in ("si", "sí"), None


def backup_base():
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(BASE_PATH)
    destino = f"{base}_respaldo_{stamp}{ext}"
    shutil.copy2(BASE_PATH, destino)
    return destino


def load_facturacion():
    wb = openpyxl.load_workbook(FACTURACION_PATH, data_only=True)
    ws = wb["FACTURAS"]
    by_nfac = defaultdict(list)
    for row_cells in ws.iter_rows(min_row=2):
        nfac = row_cells[F_NFAC].value
        if nfac is None:
            continue
        # Se agrega al final el texto del comentario de la celda 'Pagada'
        # (fr[-1]), si tiene uno — ver estado_pago_facturacion().
        comentario_pagada = row_cells[F_PAGADA].comment.text if row_cells[F_PAGADA].comment else None
        row = tuple(c.value for c in row_cells) + (comentario_pagada,)
        by_nfac[nfac].append(row)
    return by_nfac


def load_dim_clientes(wb_base):
    ws = wb_base["Dim_Clientes"]
    by_rut = defaultdict(list)
    max_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cliente_id, nombre, rut = row[0], row[1], row[2]
        if isinstance(cliente_id, str):
            m = re.match(r"^CLI-(\d+)$", cliente_id.strip())
            if m:
                max_num = max(max_num, int(m.group(1)))
        nr = norm_rut(rut)
        if nr:
            by_rut[nr].append((cliente_id, (nombre or "").strip().upper()))
    return by_rut, max_num


def resolve_cliente_id(rut_to_candidates, rut, nombre_facturacion):
    candidates = rut_to_candidates.get(norm_rut(rut))
    if not candidates:
        return None, None
    if len(candidates) == 1:
        return candidates[0][0], None
    # RUT con más de un cliente en Dim_Clientes (caso ya documentado en
    # QA_Calidad_Datos) — solo aceptar si el nombre
    # calza exactamente; si no, no adivinar.
    nombre_norm = (nombre_facturacion or "").strip().upper()
    exact = [c for c in candidates if c[1] == nombre_norm]
    if len(exact) == 1:
        return exact[0][0], None
    return None, (
        f"RUT {rut} tiene {len(candidates)} clientes distintos en Dim_Clientes "
        f"({', '.join(c[0] for c in candidates)}) — no se pudo determinar cuál corresponde a '{nombre_facturacion}'"
    )


def update_existing(ws_base, fact_by_nfac):
    changes = []
    for r in range(2, ws_base.max_row + 1):
        nfac = ws_base.cell(r, B_NFAC).value
        if nfac is None:
            continue
        candidates = fact_by_nfac.get(nfac)
        if not candidates:
            continue

        total = ws_base.cell(r, B_TOTAL).value
        same_total = [c for c in candidates if c[F_TOTAL] == total]
        fr = (same_total or candidates)[0]

        base_pagada = (ws_base.cell(r, B_PAGADA).value or "").strip()
        base_special = base_pagada in ("En demanda", "Sin registrar")
        base_paid = base_pagada.lower() in ("si", "sí")
        fact_paid, fact_fecha_pago = estado_pago_facturacion(fr)

        if base_special or base_paid or not fact_paid:
            continue

        cliente = ws_base.cell(r, B_CLIENTE_REG).value
        ws_base.cell(r, B_PAGADA).value = "Sí"
        fact_pagada_txt = (fr[F_PAGADA] or "").strip()
        if fact_pagada_txt.lower() in ("si", "sí"):
            origen = "flag Pagada"
        elif fecha_desde_comentario(fr[-1]) is not None:
            origen = "comentario en celda Pagada (nota manual de pago)"
        else:
            origen = "fecha real en Banco (flag Pagada sin marcar)"
        detail = f"ACTUALIZADA fila {r} | N°{nfac} | {cliente} | ${total:,.0f} -> Pagada=Sí (detectado por {origen})"

        fecha_actual = ws_base.cell(r, B_FECHA_REAL_PAGO).value
        if fecha_actual is None and fact_fecha_pago is not None:
            ws_base.cell(r, B_FECHA_REAL_PAGO).value = fact_fecha_pago
            detail += f", Fecha_Real_Pago={fact_fecha_pago.date()}"

        changes.append(detail)
    return changes


def ingest_new(ws_base, ws_dim, fact_by_nfac, rut_to_candidates, max_cliente_num):
    existing_nfacs = set()
    max_row = ws_base.max_row
    max_factura_id = 0
    for r in range(2, max_row + 1):
        nfac = ws_base.cell(r, B_NFAC).value
        if nfac is not None:
            existing_nfacs.add(nfac)
        fid = ws_base.cell(r, B_FACTURA_ID).value
        if isinstance(fid, (int, float)):
            max_factura_id = max(max_factura_id, int(fid))

    changes = []
    skipped = []
    next_row = max_row + 1
    next_factura_id = max_factura_id + 1
    next_dim_row = ws_dim.max_row + 1
    next_cliente_num = max_cliente_num + 1
    facturas_por_cliente = defaultdict(int)

    for nfac, rows in fact_by_nfac.items():
        if nfac in existing_nfacs:
            continue
        for fr in rows:
            fem, fvenc, neto = fr[F_FEM], fr[F_FVENC], fr[F_NETO]
            # Mismo criterio que la consulta Power Query de Fact_Facturas:
            # fecha de emisión, fecha de vencimiento y neto no vacíos/no cero.
            if not isinstance(fem, datetime.datetime):
                continue
            if not isinstance(fvenc, datetime.datetime):
                continue
            if not neto:
                continue

            rut = fr[F_RUT]
            cliente = fr[F_CLIENTE]
            total = fr[F_TOTAL]
            cliente_id, ambiguity = resolve_cliente_id(rut_to_candidates, rut, cliente)

            if not cliente_id and ambiguity:
                skipped.append(f"N°{nfac} | {cliente} | {ambiguity} — revisar a mano")
                continue

            if not cliente_id:
                # RUT no encontrado en absoluto -> crear cliente nuevo
                cliente_id = f"CLI-{next_cliente_num:04d}"
                rut_dc = rut if rut else "Sin RUT"
                ws_dim.cell(next_dim_row, DC_CLIENTE_ID).value = cliente_id
                ws_dim.cell(next_dim_row, DC_NOMBRE).value = cliente
                ws_dim.cell(next_dim_row, DC_RUT).value = rut_dc
                ws_dim.cell(next_dim_row, DC_N_OTS).value = 0
                ws_dim.cell(next_dim_row, DC_N_FACTURAS).value = 0
                changes.append(f"CLIENTE NUEVO fila {next_dim_row} | {cliente_id} | {cliente} | RUT {rut_dc}")
                nr = norm_rut(rut)
                if nr:
                    rut_to_candidates[nr].append((cliente_id, (cliente or "").strip().upper()))
                next_dim_row += 1
                next_cliente_num += 1

            neto_val = neto
            total_val = total if total is not None else neto_val
            iva_val = fr[F_IVA] if fr[F_IVA] is not None else (total_val - neto_val)
            fact_paid, fecha_real_pago = estado_pago_facturacion(fr)
            pagada = "Sí" if fact_paid else ((fr[F_PAGADA] or "No").strip() or "No")

            r = next_row
            ws_base.cell(r, B_FACTURA_ID).value = next_factura_id
            ws_base.cell(r, B_NCORR).value = nfac
            ws_base.cell(r, B_NFAC).value = nfac
            ws_base.cell(r, B_TIPODOC).value = "Factura"
            ws_base.cell(r, B_CLIENTE_ID).value = cliente_id
            ws_base.cell(r, B_CLIENTE_REG).value = cliente
            ws_base.cell(r, B_RUT_REG).value = rut
            ws_base.cell(r, B_FEM).value = fem
            ws_base.cell(r, B_FVENC).value = fvenc
            ws_base.cell(r, B_NETO).value = neto_val
            ws_base.cell(r, B_IVA).value = iva_val
            ws_base.cell(r, B_TOTAL).value = total_val
            ws_base.cell(r, B_OC).value = fr[F_OC]
            ws_base.cell(r, B_GUIA).value = fr[F_GUIA]
            ws_base.cell(r, B_CEDIDA).value = fr[F_CEDIDA]
            ws_base.cell(r, B_FACTORING).value = fr[F_FACTORING]
            ws_base.cell(r, B_PAGADA).value = "Sí" if pagada.lower() in ("si", "sí") else pagada
            ws_base.cell(r, B_FECHA_REAL_PAGO).value = fecha_real_pago

            changes.append(f"NUEVA fila {r} | N°{nfac} | {cliente} | ${total_val:,.0f}")
            facturas_por_cliente[cliente_id] += 1
            next_row += 1
            next_factura_id += 1

    # actualizar conteo N_Facturas en Dim_Clientes para clientes con facturas nuevas
    if facturas_por_cliente:
        for r in range(2, ws_dim.max_row + 1):
            cid = ws_dim.cell(r, DC_CLIENTE_ID).value
            if cid in facturas_por_cliente:
                actual = ws_dim.cell(r, DC_N_FACTURAS).value or 0
                ws_dim.cell(r, DC_N_FACTURAS).value = actual + facturas_por_cliente[cid]

    return changes, skipped


def write_qa_notes(wb_base, skipped):
    if not skipped:
        return
    ws_qa = wb_base["QA_Calidad_Datos"]
    stamp = datetime.datetime.now().strftime("%d-%m-%Y")
    r = ws_qa.max_row + 2
    ws_qa.cell(r, 2).value = f"Sync automático {stamp} — facturas nuevas que requieren revisión manual"
    r += 1
    for s in skipped:
        ws_qa.cell(r, 1).value = "REVISAR"
        ws_qa.cell(r, 2).value = s
        r += 1


def main():
    try:
        fact_by_nfac = load_facturacion()
    except PermissionError:
        log("Facturacion clientes JVA.xlsx está abierto en Excel — se omite este ciclo.")
        return
    except Exception as e:
        log(f"ERROR abriendo Facturacion clientes JVA.xlsx: {e}")
        return

    try:
        wb_base = openpyxl.load_workbook(BASE_PATH)
    except PermissionError:
        log("BASE_DE_DATOS_JVA.xlsx está abierto (Excel o Power BI Desktop con lock) — se omite este ciclo.")
        return
    except Exception as e:
        log(f"ERROR abriendo BASE_DE_DATOS_JVA.xlsx: {e}")
        return

    ws_base = wb_base["Fact_Facturas"]
    ws_dim = wb_base["Dim_Clientes"]
    rut_to_candidates, max_cliente_num = load_dim_clientes(wb_base)

    updated = update_existing(ws_base, fact_by_nfac)
    inserted, skipped = ingest_new(ws_base, ws_dim, fact_by_nfac, rut_to_candidates, max_cliente_num)

    for s in skipped:
        log("OMITIDA (requiere revisión manual): " + s)
    write_qa_notes(wb_base, skipped)

    all_changes = updated + inserted
    if not all_changes:
        log("Sin cambios — Base de Datos ya está sincronizada con Facturacion.")
        return

    try:
        respaldo = backup_base()
    except Exception as e:
        log(f"ERROR creando respaldo, no se aplican cambios por seguridad: {e}")
        return

    try:
        wb_base.save(BASE_PATH)
    except PermissionError:
        log(
            f"Se detectaron {len(all_changes)} cambios pero el archivo está bloqueado para guardar — se reintenta el próximo ciclo. Respaldo sin usar: {respaldo}"
        )
        return

    log(f"Respaldo creado: {respaldo}")
    log(f"{len(updated)} factura(s) actualizada(s), {len(inserted)} factura(s)/cliente(s) nuevo(s):")
    for c in all_changes:
        log("  - " + c)


if __name__ == "__main__":
    sys.exit(main())
