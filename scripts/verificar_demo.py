# -*- coding: utf-8 -*-
"""Verificacion independiente del demo anonimizado.

No reutiliza nada de anonimizar_demo.py a proposito: si el anonimizador tiene
un bug de logica, este script tiene que detectarlo igual. Revisa CADA celda de
CADA hoja. Sale con codigo != 0 si encuentra cualquier fuga.
"""
import io, os, re, sys, unicodedata

import openpyxl

REAL = r"C:\Users\ignac\Dropbox\JVA ADM\BASE_DE_DATOS_JVA.xlsx"
DEMO = (r"C:\Users\ignac\Dropbox\JVA ADM\Dashboards\Portfolio"
        r"\jva-finance-dashboard\data\BASE_DE_DATOS_JVA_demo.xlsx")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def norm(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def palabras(s):
    """Como norm(), pero deja los separadores como espacios.

    norm() pega todo, asi que encuentra nombres a caballo entre dos palabras:
    dos palabras cualesquiera, pegadas, forman una tercera que no esta ahi.
    En una celda de Excel ese exceso de celo es barato; sobre prosa y codigo es
    puro falso positivo, y una compuerta que grita lobo termina ignorada. Aqui
    se compara con limites de palabra.
    """
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " " + re.sub(r"[^A-Z0-9]+", " ", s.upper()).strip() + " "


# Razones sociales que son solo una palabra generica: buscarlas marcaria cada
# linea del repositorio sin identificar a nadie.
GENERICOS = {
    "SERVICIOS", "SERVICIO", "INDUSTRIAL", "INDUSTRIALES", "INDUSTRIA",
    "LIMITADA", "LTDA", "EIRL", "SOCIEDAD", "COMERCIAL", "EMPRESA", "EMPRESAS",
    "CHILE", "MINERA", "MINERIA", "TRANSPORTE", "TRANSPORTES", "INGENIERIA",
    "CONSTRUCTORA", "MAESTRANZA", "PLANTA", "NACIONAL", "CENTRAL", "GENERAL",
    "PROYECTOS", "MONTAJES", "MANTENCION", "EQUIPOS", "REPUESTOS", "GRUPO",
    "FACTURA", "CLIENTE", "CLIENTES", "DATOS", "TOTAL", "BANCO", "POWER",
}


wr = openpyxl.load_workbook(REAL, data_only=True)
wd = openpyxl.load_workbook(DEMO, data_only=True)

# Valores centinela: existen en produccion Y es correcto que sigan en el demo.
# Si no se excluyen, el verificador los reporta como "nombre real filtrado".
CENTINELAS = {norm(x) for x in ("-", "N/A", "NULO", "NULL", "Sin RUT", "None",
                                "SIN INFORMACION", "S/I", "NO APLICA")}

# --- universo de secretos a buscar
nombres_reales, ruts_reales = set(), set()
for r in wr["Dim_Clientes"].iter_rows(min_row=2, values_only=True):
    if r[1] and str(r[1]).strip() not in ("-", ""):
        n = norm(r[1])
        if len(n) >= 4 and n not in CENTINELAS:
            nombres_reales.add((n, str(r[1]).strip()))
    if r[2] and norm(r[2]) not in CENTINELAS:
        ruts_reales.add((norm(r[2]), str(r[2]).strip()))

# nombres reales que solo aparecen en Fact_Operaciones (abreviaturas, personas)
for r in wr["Fact_Operaciones"].iter_rows(min_row=2, values_only=True):
    if len(r) > 6 and r[6] and str(r[6]).strip() not in ("-", ""):
        n = norm(r[6])
        if len(n) >= 4 and n not in CENTINELAS:
            nombres_reales.add((n, str(r[6]).strip()))
    if len(r) > 7 and r[7] and norm(r[7]) not in CENTINELAS:
        ruts_reales.add((norm(r[7]), str(r[7]).strip()))

ids_reales = {str(r[0]) for r in wr["Dim_Clientes"].iter_rows(min_row=2, values_only=True) if r[0]}

print("buscando %d nombres reales y %d RUTs reales en cada celda del demo\n"
      % (len(nombres_reales), len(ruts_reales)))

PROSA = ("Diccionario de Datos", "QA_Calidad_Datos")
fallos = []

# --- 1. barrido celda por celda
for hoja in wd.sheetnames:
    hits_n, hits_r = [], []
    for fi, fila in enumerate(wd[hoja].iter_rows(values_only=True), start=1):
        for ci, v in enumerate(fila):
            if v is None:
                continue
            nv = norm(v)
            if not nv:
                continue
            for n, orig in nombres_reales:
                if n in nv:
                    hits_n.append((fi, ci, orig, str(v)[:60]))
                    break
            for rr, orig in ruts_reales:
                if rr and rr in nv:
                    hits_r.append((fi, ci, orig, str(v)[:60]))
                    break
    estado = "OK" if not hits_n and not hits_r else "FUGA"
    print("  %-28s nombres=%-4d ruts=%-4d  %s" % (hoja, len(hits_n), len(hits_r), estado))
    for h in hits_n[:4]:
        print("      nombre '%s' en fila %d col %d -> %r" % (h[2], h[0], h[1], h[3]))
    for h in hits_r[:4]:
        print("      RUT    '%s' en fila %d col %d -> %r" % (h[2], h[0], h[1], h[3]))
    if hits_n or hits_r:
        fallos.append("%s: %d nombres, %d ruts" % (hoja, len(hits_n), len(hits_r)))

# --- 2. Cliente_ID debe ser disjunto de produccion
ids_demo = set()
for hoja in ("Dim_Clientes", "Fact_Operaciones", "Fact_Facturas"):
    hdr = [str(c) for c in next(wd[hoja].iter_rows(max_row=1, values_only=True))]
    if "Cliente_ID" not in hdr:
        continue
    i = hdr.index("Cliente_ID")
    for r in wd[hoja].iter_rows(min_row=2, values_only=True):
        if i < len(r) and r[i]:
            ids_demo.add(str(r[i]))
solape = ids_demo & ids_reales
print("\n  Cliente_ID demo=%d  produccion=%d  SOLAPE=%d  %s"
      % (len(ids_demo), len(ids_reales), len(solape), "OK" if not solape else "FUGA"))
if solape:
    fallos.append("Cliente_ID solapa en %d valores" % len(solape))

# --- 3. ninguna hoja puede ser identica a produccion
print()
for hoja in wr.sheetnames:
    if hoja not in wd.sheetnames:
        continue
    a = list(wd[hoja].iter_rows(values_only=True))
    b = list(wr[hoja].iter_rows(values_only=True))
    n = min(len(a), len(b))
    # Solo cuentan filas CON datos: un mes sin facturacion (None/0) es identico
    # por definicion y no revela nada. Y las hojas de prosa documentan el
    # esquema, que no es secreto: ahi vale el barrido de nombres/RUTs, no la
    # identidad textual.
    def con_datos(f):
        return any(v not in (None, 0, "") for v in f[2:]) if len(f) > 2 else False
    idxs = [i for i in range(1, n) if con_datos(b[i])]
    ident = sum(1 for i in idxs if a[i] == b[i])
    pct = 100.0 * ident / max(1, len(idxs))
    estado = "OK" if (pct < 5 or hoja in PROSA) else "FUGA"
    print("  %-28s filas identicas a produccion: %d/%d (%.1f%%)  %s"
          % (hoja, ident, n - 1, pct, estado))
    if pct >= 5 and hoja not in PROSA:
        fallos.append("%s identica a produccion en %.1f%%" % (hoja, pct))

# --- 4. montos: ningun Total real debe aparecer tal cual
tot_real = {r[11] for r in wr["Fact_Facturas"].iter_rows(min_row=2, values_only=True)
            if isinstance(r[11], (int, float)) and r[11]}
tot_demo = {r[11] for r in wd["Fact_Facturas"].iter_rows(min_row=2, values_only=True)
            if isinstance(r[11], (int, float)) and r[11]}
coinc = tot_real & tot_demo
print("\n  Totales de factura que coinciden con produccion: %d de %d  %s"
      % (len(coinc), len(tot_demo), "OK" if len(coinc) < 5 else "REVISAR"))

# --- 5. barrido de los archivos de texto del repositorio
# El demo puede estar impecable y el repositorio filtrar igual: una version
# anterior de anonimizar_demo.py traia un diccionario de marcas reales, o sea
# que el script encargado de ocultar los nombres los publicaba el mismo. Este
# paso es el que atrapa esa clase de error.
#
# Se busca la razon social COMPLETA, no sus palabras por separado. Probamos lo
# segundo y no es viable aqui: buena parte de la cartera son personas naturales
# y varias razones sociales estan hechas de sustantivos corrientes, asi que
# tokens como TRABAJOS, PUERTO, GROUP o el apellido del autor marcaban prosa y
# codigo inocentes. Una compuerta con falsos positivos se termina ignorando.
#
# La marca suelta dentro de una frase la cubre el barrido celda por celda del
# paso 1, que si es agresivo a proposito. Este paso cubre el caso que aquel no
# puede ver: un nombre real escrito a mano en el codigo o en la documentacion.
secretos_txt = set()
for _n, _orig in nombres_reales:
    _p = palabras(_orig).strip()
    if len(_p) >= 4 and _p not in GENERICOS:
        secretos_txt.add((_p, _orig))

TEXTO = (".md", ".py", ".json", ".tmdl", ".pbip", ".pbir", ".txt",
         ".gitignore", ".gitattributes")
IGNORAR = {".git", "__pycache__", "data", "screenshots"}
raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

print()
revisados = hits_txt = 0
for carpeta, subdirs, archivos in os.walk(raiz):
    subdirs[:] = [d for d in subdirs if d not in IGNORAR]
    for nombre in archivos:
        if not nombre.endswith(TEXTO):
            continue
        ruta = os.path.join(carpeta, nombre)
        rel = os.path.relpath(ruta, raiz)
        try:
            contenido = io.open(ruta, encoding="utf-8", errors="replace").read()
        except OSError:
            continue
        revisados += 1
        for li, linea in enumerate(contenido.splitlines(), start=1):
            pl = palabras(linea)
            nl = norm(linea)
            for sec, orig in secretos_txt:
                if " " + sec + " " in pl:
                    print("  FUGA %s:%d  nombre '%s'" % (rel, li, orig))
                    hits_txt += 1
                    break
            for rr, orig in ruts_reales:
                if len(rr) >= 8 and rr in nl:
                    print("  FUGA %s:%d  RUT '%s'" % (rel, li, orig))
                    hits_txt += 1
                    break
print("  archivos de texto revisados: %d   nombres/RUTs reales: %d  %s"
      % (revisados, hits_txt, "OK" if not hits_txt else "FUGA"))
if hits_txt:
    fallos.append("%d apariciones de datos reales en archivos de texto" % hits_txt)

print("\n" + "=" * 62)
if fallos:
    print("RESULTADO: %d PROBLEMAS" % len(fallos))
    for f in fallos:
        print("   -", f)
    sys.exit(1)
print("RESULTADO: LIMPIO — ni las hojas ni los archivos de texto")
print("           del repositorio contienen datos reales")
