# -*- coding: utf-8 -*-
"""Genera BASE_DE_DATOS_JVA_demo.xlsx anonimizando LAS SEIS hojas.

v2. La v1 mapeaba por texto exacto del nombre en Dim_Clientes, y las variantes
tipeadas a mano en Fact_Operaciones/Fact_Facturas (abreviaturas, siglas,
espaciado distinto alrededor del '&') no calzaban y pasaban intactas — el mismo tipo
de fallo que causo la fuga original. Ahora el mapeo es POR Cliente_ID de la
fila, no por texto: si la fila tiene ID, su nombre y RUT se reescriben si o si.

Decisiones:
  - Cliente_ID: espacio de llaves nuevo y barajado, disjunto de produccion.
  - Nombre/RUT: derivados del Cliente_ID de la fila. Las variantes de tipeo se
    preservan como variantes del nombre FICTICIO (asi el dashboard sigue
    mostrando el mismo problema de calidad de datos, sin exponer a nadie).
  - Montos: escalados por un factor unico. Las proporciones — y por tanto todos
    los porcentajes del dashboard — se conservan exactas.
  - Hojas de prosa: 'QA_Calidad_Datos' se reemplaza por una nota generica (218
    filas de texto sobre clientes y personas reales no se sanean confiablemente
    a parches); 'Diccionario de Datos' conserva la documentacion de esquema y
    se le sustituyen los ejemplos.
"""
import openpyxl, re, unicodedata, os

REAL = r"C:\Users\ignac\Dropbox\JVA ADM\BASE_DE_DATOS_JVA.xlsx"
OUT = (r"C:\Users\ignac\Dropbox\JVA ADM\Dashboards\Portfolio"
       r"\jva-finance-dashboard\data\BASE_DE_DATOS_JVA_demo.xlsx")

ESCALA = 0.7314
SEMILLA = 20260901
SENTINELAS = {"", "-", "N/A", "NULO", "NULL", "SIN RUT", "SIN INFORMACION", "NONE"}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def dv_rut(cuerpo):
    s, m = 0, 2
    for d in reversed(str(cuerpo)):
        s += int(d) * m
        m = 2 if m == 7 else m + 1
    r = 11 - (s % 11)
    return {11: "0", 10: "K"}.get(r, str(r))


def fmt_rut(cuerpo):
    c, miles = str(cuerpo), ""
    while len(c) > 3:
        miles, c = "." + c[-3:] + miles, c[:-3]
    return c + miles + "-" + dv_rut(cuerpo)


class LCG:
    def __init__(self, seed): self.s = seed
    def next(self):
        self.s = (self.s * 1103515245 + 12345) % (2 ** 31)
        return self.s
    def below(self, n): return self.next() % n
    def shuffle(self, l):
        for i in range(len(l) - 1, 0, -1):
            j = self.below(i + 1)
            l[i], l[j] = l[j], l[i]
        return l


rng = LCG(SEMILLA)

GIRO = ["MAESTRANZA", "SERVICIOS INDUSTRIALES", "INGENIERIA Y MONTAJES",
        "METALMECANICA", "SOLUCIONES INDUSTRIALES", "COMERCIAL TECNICA",
        "MANTENIMIENTO INDUSTRIAL", "AUTOMATIZACION", "HIDRAULICA",
        "TRANSPORTES Y LOGISTICA", "CONSTRUCTORA", "SUMINISTROS TECNICOS",
        "BOMBAS Y VALVULAS", "MONTAJES ELECTRICOS", "FUNDICION"]
GEO = ["ANDINA", "PACIFICO", "DEL NORTE", "CORDILLERA", "ALTIPLANO", "COSTERA",
       "AUSTRAL", "CENTRAL", "PAMPA", "SALAR", "PUERTO", "VALLE", "MERIDIANA",
       "ATACAMA", "ARIDA", "BOREAL", "ORIENTE", "TERRAL", "SERENA", "LOA"]
FORMA = ["SPA", "LTDA", "S.A.", "E.I.R.L.", "SPA", "LTDA"]
PILA = ["ANDRES", "BEATRIZ", "CAMILO", "DANIELA", "ESTEBAN", "FERNANDA",
        "GONZALO", "HELENA", "IGNACIO", "JAVIERA", "LORENZO", "MARISOL",
        "NICOLAS", "OLIVIA", "PATRICIO", "ROSARIO", "SEBASTIAN", "TAMARA"]
APE = ["ARANEDA", "BUSTOS", "CARRASCO", "DONOSO", "ESPINOZA", "FUENZALIDA",
       "GALLARDO", "HERRERA", "IBACACHE", "JARA", "LAGOS", "MELLADO",
       "NAVARRETE", "ORELLANA", "PAREDES", "QUINTEROS", "RIVEROS", "SOTO"]


def es_persona(n):
    n = str(n).upper()
    if any(f in n for f in ("SPA", "LTDA", "S.A", "E.I.R.L", "SCM", "INDUSTRIAL",
                            "SERVICIOS", "COMERCIAL", "MAESTRANZA", "INGENIERIA",
                            "SOCIEDAD", "TRANSPORTE", "CONSTRUC")):
        return False
    return len(n.split()) in (2, 3)


wb = openpyxl.load_workbook(REAL)

# ------------------------------------------------------- 1. mapeos por Cliente_ID
reales = [(str(r[0]), str(r[1] or "").strip())
          for r in wb["Dim_Clientes"].iter_rows(min_row=2, values_only=True) if r[0]]

ids_nuevos = rng.shuffle(["C-%04d" % n for n in range(1000, 1000 + len(reales))])
MAP_ID = {rid: ids_nuevos[i] for i, (rid, _) in enumerate(reales)}

cuerpos = rng.shuffle([str(n) for n in range(61000000, 61000000 + len(reales) * 4)])
NOMBRE_POR_ID, RUT_POR_ID, usados = {}, {}, set()
# Token distintivo PRIMERO. Si el nombre empieza por el giro ("INGENIERIA Y
# MONTAJES ..."), al truncarse en un grafico de barras todos se ven iguales:
# cuatro barras que dicen "INGENIERIA Y MO..." no comunican nada.
PREF = ["AND", "VER", "NOR", "AUS", "COR", "PAM", "SAL", "TER", "BOR", "MER",
        "LOA", "ATA", "QUIL", "RAN", "MAI", "TAL", "CUR", "LIN", "VAL", "PUR",
        "CAL", "HUE", "MEL", "TOC", "ILL", "CHA", "PEL", "REN", "VIC", "ZAP"]
SUF = ["TEX", "MAX", "VOR", "SUR", "MEC", "TON", "LAB", "GEN", "FIN", "NOR",
        "PAR", "VIA", "ZAN", "CO", "SAN", "MET"]
tokens = rng.shuffle([a + b for a in PREF for b in SUF])
usados_tok = set()

for i, (rid, rnom) in enumerate(reales):
    if es_persona(rnom):
        for _ in range(120):
            cand = "%s %s" % (PILA[rng.below(len(PILA))], APE[rng.below(len(APE))])
            if cand not in usados:
                break
    else:
        tok = tokens[len(usados_tok) % len(tokens)]
        usados_tok.add(tok)
        cand = "%s %s %s" % (tok, GIRO[rng.below(len(GIRO))], FORMA[rng.below(len(FORMA))])
    usados.add(cand)
    NOMBRE_POR_ID[rid] = cand
    RUT_POR_ID[rid] = fmt_rut(cuerpos[i])

# variantes de tipeo -> variante del nombre FICTICIO (preserva el problema de
# calidad de datos sin exponer al cliente real)
def variante(nombre_falso, texto_real):
    t = str(texto_real).upper()
    if "." in t and "." not in nombre_falso:
        return nombre_falso.replace(" ", ". ", 1)
    if len(t) <= 12:
        return nombre_falso.split()[0]
    if " & " in t:
        return nombre_falso.replace(" ", " & ", 1)
    return nombre_falso


RUT_RE = re.compile(r"\b\d{1,2}[.\s]?\d{3}[.\s]?\d{3}\s*[-\u2013]?\s*[\dkK]\b")

# Sustitucion de nombres reales en TEXTO LIBRE.
#
# El mapa NO se escribe a mano: se deriva de los 213 clientes reales. Un mapa
# manual deja siempre residuos -- marcas sueltas dentro de una descripcion de
# equipo, siglas, razones sociales sin su forma legal --, que es exactamente
# como sobrevivio la fuga original. Aca cualquier nombre real que aparezca en
# una descripcion de equipo, un motivo de ingreso o una nota queda cubierto.
#
# NO escribir aqui una lista de clientes a mano. Una version anterior de este
# script traia un diccionario de marcas reales para que las sustituciones
# quedaran mas prolijas, y el efecto fue que el script encargado de ocultar los
# nombres los publicaba el mismo. El mapa se deriva enteramente del Excel de
# produccion, que no forma parte de este repositorio.

_SUST = {}
_PALABRA_COMUN = {
    "SPA", "LTDA", "S.A.", "SA", "EIRL", "E.I.R.L.", "SCM", "Y", "DE", "DEL",
    "LA", "EL", "LOS", "LAS", "SERVICIOS", "INDUSTRIAL", "INDUSTRIALES",
    "COMERCIAL", "SOCIEDAD", "INGENIERIA", "LIMITADA", "CHILE", "S.A",
    "MAESTRANZA", "TRANSPORTES", "CONSTRUCTORA", "MANTENIMIENTO", "NULO",
}
for _rid, _rnom in reales:
    if not _rnom or _rnom.upper() in SENTINELAS:
        continue
    _falso = NOMBRE_POR_ID.get(_rid)
    if not _falso:
        continue
    _SUST.setdefault(_rnom.upper(), _falso)
    # tambien el nombre sin la forma societaria: la prosa suele citar
    # la marca sin sufijo cuando la razon social lo incluye ("X" vs "X SPA")
    _base = re.sub(r"(?i)[\s,.]*(SPA|LTDA|LIMITADA|S\.?A\.?|E\.?I\.?R\.?L\.?|SCM)[\s.]*$",
                   "", _rnom.upper()).strip()
    if len(_base) >= 4 and _base not in _PALABRA_COMUN:
        _SUST.setdefault(_base, _falso)
    # token distintivo de la razon social, para atrapar la marca
    # suelta dentro de descripciones de equipo
    for _tok in re.split(r"[^A-Za-z0-9À-ſ]+", _rnom.upper()):
        if len(_tok) >= 5 and _tok not in _PALABRA_COMUN:
            _SUST.setdefault(_tok, _falso.split()[0])

_MARCA_RE = re.compile(r"(?i)\b(" + "|".join(
    re.escape(k) for k in sorted(_SUST, key=len, reverse=True)) + r")\b")


def limpiar_marcas(t):
    return _MARCA_RE.sub(lambda m: _SUST.get(m.group(1).upper(), "[cliente]"), t)


# ------------------------------------------------------- 2. hojas tabulares
def col_idx(ws):
    hdr = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    return hdr, {h: i for i, h in enumerate(hdr)}


for hoja in ("Dim_Clientes", "Fact_Operaciones", "Fact_Facturas"):
    ws = wb[hoja]
    hdr, ix = col_idx(ws)
    n = 0
    for fila in ws.iter_rows(min_row=2):
        cid = None
        if "Cliente_ID" in ix and fila[ix["Cliente_ID"]].value:
            cid = str(fila[ix["Cliente_ID"]].value)
        for nombre_col, celda in zip(hdr, fila):
            v = celda.value
            if v is None:
                continue
            c = nombre_col.lower()
            if c == "cliente_id" and cid:
                celda.value = MAP_ID.get(cid, "C-9999"); n += 1
            elif c in ("nombre_cliente", "cliente_registrado"):
                if str(v).strip().upper() in SENTINELAS:
                    continue
                base = NOMBRE_POR_ID.get(cid, "CLIENTE ANONIMO SPA")
                celda.value = base if c == "nombre_cliente" else variante(base, v); n += 1
            elif "rut" in c:
                if str(v).strip().upper() in SENTINELAS:
                    continue
                celda.value = RUT_POR_ID.get(cid, fmt_rut("61999999")); n += 1
            elif c in ("neto", "iva", "total", "ventas", "compras", "ppm_5pct", "ppm_voluntario"):
                if isinstance(v, (int, float)):
                    celda.value = round(v * ESCALA); n += 1
            elif c in ("orden_compra", "guia_despacho", "equipo_componente",
                       "motivo_ingreso", "motivo_ingreso_original", "estado_nota",
                       "fecha_despacho_nota"):
                if isinstance(v, str):
                    nv = limpiar_marcas(RUT_RE.sub("[dato omitido]", v))
                    if nv != v:
                        celda.value = nv; n += 1
    print("  %-28s celdas modificadas: %d" % (hoja, n))

# Fact_VentasCompras_Mensual: escalar montos
ws = wb["Fact_VentasCompras_Mensual"]
hdr, ix = col_idx(ws)
n = 0
for fila in ws.iter_rows(min_row=2):
    for nombre_col, celda in zip(hdr, fila):
        if nombre_col.lower() in ("ventas", "compras", "ppm_5pct", "ppm_voluntario"):
            if isinstance(celda.value, (int, float)) and celda.value:
                celda.value = round(celda.value * ESCALA); n += 1
print("  %-28s celdas modificadas: %d" % ("Fact_VentasCompras_Mensual", n))

# IVA / Total recalculados tras el escalado
ws = wb["Fact_Facturas"]
hdr, ix = col_idx(ws)
for fila in ws.iter_rows(min_row=2):
    v = fila[ix["Neto"]].value
    if isinstance(v, (int, float)):
        fila[ix["IVA"]].value = round(v * 0.19)
        fila[ix["Total"]].value = round(v * 1.19)

# ------------------------------------------------------- 3. hojas de prosa
del wb["QA_Calidad_Datos"]
qa = wb.create_sheet("QA_Calidad_Datos")
for i, linea in enumerate([
    "CALIDAD DE DATOS - REVISION MANUAL",
    "",
    "El archivo de produccion incluye aqui una bitacora de revision de calidad de datos:",
    "clientes duplicados por RUT compartido, razones sociales tipeadas de varias formas,",
    "ordenes de trabajo sin RUT valido y facturas con estado de pago inconsistente.",
    "",
    "Esa bitacora nombra clientes y personas reales, por lo que NO se incluye en esta",
    "version de demostracion. El script scripts/sync_facturacion_a_base.py escribe aqui",
    "sus hallazgos automaticos en cada corrida.",
    "",
    "Los problemas de calidad que documenta estan descritos, ya anonimizados, en el README.",
], start=1):
    qa.cell(row=i, column=1, value=linea)

# Diccionario de Datos: conservar el esquema, sustituir ejemplos con clientes reales
ws = wb["Diccionario de Datos"]
NOMBRES_REALES = sorted({n for _, n in reales if n and n.upper() not in SENTINELAS},
                        key=len, reverse=True)
n = 0
for fila in ws.iter_rows():
    for celda in fila:
        if not isinstance(celda.value, str):
            continue
        v = celda.value
        for real in NOMBRES_REALES:
            if len(real) < 4:
                continue
            while real.upper() in v.upper():
                i = v.upper().find(real.upper())
                v = v[:i] + "[cliente]" + v[i + len(real):]
        v = limpiar_marcas(RUT_RE.sub("[RUT]", v))
        # abreviaturas y siglas sueltas que quedan en los ejemplos
        v = re.sub(r"'[A-Z][A-Z0-9&. ]{1,24}'", "'[ejemplo]'", v)
        if v != celda.value:
            celda.value = v; n += 1
print("  %-28s celdas modificadas: %d" % ("Diccionario de Datos", n))
print("  %-28s reemplazada por nota generica" % "QA_Calidad_Datos")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("\nescrito:", OUT)
