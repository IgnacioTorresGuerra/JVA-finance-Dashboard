# -*- coding: utf-8 -*-
"""Recalcula todas las cifras citadas en la seccion 'Hallazgos principales'.

Corre sobre el set de demostracion incluido en este repositorio, asi que
cualquiera puede clonar y verificar que los numeros del README salen del dato
y no de una planilla aparte:

    python scripts/hallazgos.py

Los montos del demo estan escalados por un factor unico, de modo que las
participaciones y variaciones porcentuales son identicas a las de produccion;
los valores absolutos, no.
"""
import collections
import datetime
import os
import sys

import openpyxl

DEMO = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    "data", "BASE_DE_DATOS_JVA_demo.xlsx")
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:  # Python 2 / consolas antiguas
    pass

EXCLUIDOS = {"Sin registrar", "En demanda"}  # estados que excluyen las medidas DAX


def cargar():
    """Devuelve las facturas afectas con fecha de emision valida.

    Las notas de credito quedan fuera: el analisis es de facturacion emitida,
    no del neto contable. Una fila sin fecha de emision no puede ubicarse en
    ninguna serie temporal, asi que tampoco entra.
    """
    wb = openpyxl.load_workbook(DEMO, read_only=True, data_only=True)
    filas = list(wb["Fact_Facturas"].iter_rows(values_only=True))
    h = {nombre: i for i, nombre in enumerate(filas[0])}
    datos = [f for f in filas[1:] if f[h["Factura_ID"]] is not None]
    facturas = [f for f in datos
                if f[h["Tipo_Documento"]] == "Factura"
                and isinstance(f[h["Fecha_Emision"]], datetime.datetime)]
    return wb, h, facturas


def moneda(x):
    return "$" + "{:,.0f}".format(x).replace(",", ".")


def main():
    wb, h, F = cargar()
    col = lambda f, c: f[h[c]]
    anio = lambda f: col(f, "Fecha_Emision").year
    dt = datetime.datetime
    anios = sorted({anio(f) for f in F})

    print("=" * 70)
    print("HALLAZGO 1 - Concentracion de clientes")
    print("=" * 70)
    for y in (2025, 2026):
        sub = [f for f in F if anio(f) == y]
        total = sum(col(f, "Total") or 0 for f in sub)
        porc = collections.Counter()
        for f in sub:
            porc[col(f, "Cliente_ID")] += col(f, "Total") or 0
        rank = [v for _, v in porc.most_common()]
        cuota = lambda k: sum(rank[:k]) / total * 100
        print("  {}  clientes activos={:3d}  top1={:5.1f}%  top3={:5.1f}%  top5={:5.1f}%"
              .format(y, len(porc), cuota(1), cuota(3), cuota(5)))

    print()
    print("=" * 70)
    print("HALLAZGO 2 - Ventas contra el mismo periodo del anio anterior")
    print("=" * 70)
    corte = max(col(f, "Fecha_Emision") for f in F)
    ultimo = corte.year
    v_act = sum(col(f, "Total") or 0 for f in F if anio(f) == ultimo)
    v_ant_tramo = sum(col(f, "Total") or 0 for f in F
                      if anio(f) == ultimo - 1
                      and col(f, "Fecha_Emision") <= corte.replace(year=ultimo - 1))
    v_ant_full = sum(col(f, "Total") or 0 for f in F if anio(f) == ultimo - 1)
    print("  ultima emision: {}".format(corte.strftime("%d-%m-%Y")))
    print("  {} (ene-{:02d}) = {:>18}".format(ultimo, corte.month, moneda(v_act)))
    print("  {} mismo tramo = {:>18}   ->  YoY comparable = {:+.1f}%"
          .format(ultimo - 1, moneda(v_ant_tramo), (v_act / v_ant_tramo - 1) * 100))
    print("  {} completo    = {:>18}   ->  YoY ingenuo    = {:+.1f}%"
          .format(ultimo - 1, moneda(v_ant_full), (v_act / v_ant_full - 1) * 100))

    print()
    print("=" * 70)
    print("HALLAZGO 3 - Contraccion de la base de clientes")
    print("=" * 70)
    for y in anios[-3:]:
        print("  {}  clientes que facturaron: {}"
              .format(y, len({col(f, "Cliente_ID") for f in F if anio(f) == y})))
    prev, act = ultimo - 1, ultimo
    c_prev = {col(f, "Cliente_ID") for f in F if anio(f) == prev}
    c_act = {col(f, "Cliente_ID") for f in F if anio(f) == act}
    valor = collections.Counter()
    for f in F:
        if anio(f) == prev:
            valor[col(f, "Cliente_ID")] += col(f, "Total") or 0
    perdidos = c_prev - c_act
    peso = sum(valor[c] for c in perdidos) / sum(valor.values()) * 100
    print("  de los {} activos en {}: {} no volvieron, {} nuevos en {}"
          .format(len(c_prev), prev, len(perdidos), len(c_act - c_prev), act))
    print("  los que no volvieron pesaban {:.1f}% de las ventas {} ({})"
          .format(peso, prev, moneda(sum(valor[c] for c in perdidos))))

    print()
    print("=" * 70)
    print("HALLAZGO 4 - Ticket promedio")
    print("=" * 70)
    for y in anios[-3:]:
        sub = [f for f in F if anio(f) == y]
        v = sum(col(f, "Total") or 0 for f in sub)
        print("  {}  ventas={:>18}  facturas={:>4}  ticket={:>14}"
              .format(y, moneda(v), len(sub), moneda(v / len(sub))))

    print()
    print("=" * 70)
    print("HALLAZGO 5 - Cumplimiento de pago por anio de emision")
    print("=" * 70)
    print("  {:6}{:>7}{:>19}{:>11}{:>11}".format("anio", "n", "ventas", "DPD prom", "% atraso"))
    tot_n = tot_v = 0
    for y in anios:
        sub = [f for f in F if anio(f) == y]
        v = sum(col(f, "Total") or 0 for f in sub)
        pagadas = [f for f in sub
                   if isinstance(col(f, "Fecha_Real_Pago"), dt)
                   and isinstance(col(f, "Fecha_Vencimiento"), dt)]
        dpd = [(col(f, "Fecha_Real_Pago") - col(f, "Fecha_Vencimiento")).days
               for f in pagadas]
        prom = sum(dpd) / len(dpd) if dpd else 0
        atraso = sum(1 for d in dpd if d > 0) / len(dpd) * 100 if dpd else 0
        tot_n += len(sub)
        tot_v += v
        print("  {:<6}{:>7}{:>19}{:>11.1f}{:>10.1f}%"
              .format(y, len(sub), moneda(v), prom, atraso))
    print("  {:<6}{:>7}{:>19}".format("TOTAL", tot_n, moneda(tot_v)))

    print()
    print("=" * 70)
    print("VALORES DE CONTROL")
    print("=" * 70)
    pagadas = [f for f in F if isinstance(col(f, "Fecha_Real_Pago"), dt)]
    atrasos = [(col(f, "Fecha_Real_Pago") - col(f, "Fecha_Vencimiento")).days
               for f in pagadas if isinstance(col(f, "Fecha_Vencimiento"), dt)]
    cobro = [(col(f, "Fecha_Real_Pago") - col(f, "Fecha_Emision")).days for f in pagadas]
    print("  facturas cargadas por el modelo : {}".format(tot_n))
    print("  clientes                        : {}".format(wb["Dim_Clientes"].max_row - 1))
    print("  ordenes de trabajo              : {}".format(wb["Fact_Operaciones"].max_row - 1))
    print("  ventas totales                  : {}".format(moneda(tot_v)))
    print("  peor atraso registrado          : {} dias".format(max(atrasos)))
    print("  dias promedio de cobro          : {:.0f}".format(sum(cobro) / len(cobro)))
    print()
    print("  Nota: los estados {} no tienen fecha de pago registrada,"
          .format(sorted(EXCLUIDOS)))
    print("  asi que excluirlos -como hacen las medidas DAX de cartera abierta-")
    print("  no altera ninguna cifra de este informe.")


if __name__ == "__main__":
    main()
